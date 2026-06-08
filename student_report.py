from openpyxl import load_workbook

FILE_PATH = ("C:/Users/HP/Desktop/python-automation-internship/week 1/student.xlsx")


def read_students():

    students = []

    try:
        wb = load_workbook(FILE_PATH)
        ws = wb["Sheet1"]

        for row in ws.iter_rows(min_row=2, values_only=True):

            name, marks = row

            if name is None or marks is None:
                continue

            students.append({
                "name": name,
                "marks": marks
            })

        if len(students) == 0:
            raise ValueError("Excel file is empty.")

        return students

    except FileNotFoundError:
        print("File not found.")
        return []

    except ValueError as e:
        print(e)
        return []

    except Exception as e:
        print("Invalid data in Excel:", e)
        return []


def generate_statistics(students):

    highest = students[0]
    lowest = students[0]

    for student in students:

        if student["marks"] > highest["marks"]:
            highest = student

        if student["marks"] < lowest["marks"]:
            lowest = student

    total = 0

    for student in students:
        total += student["marks"]

    average = total / len(students)

    above_75 = []
    below_35 = []

    pass_count = 0
    fail_count = 0

    for student in students:

        if student["marks"] > 75:
            above_75.append(student)

        if student["marks"] < 35:
            below_35.append(student)

        if student["marks"] >= 35:
            pass_count += 1
        else:
            fail_count += 1

    stats = {
        "Highest Scorer": highest["name"],
        "Lowest Scorer": lowest["name"],
        "Average": round(average, 2),
        "Pass Count": pass_count,
        "Fail Count": fail_count
    }

    return stats, above_75, below_35


def search_student(students):

    name = input("\nEnter student name to search: ")

    found = False

    for student in students:

        if student["name"].lower() == name.lower():

            print("\nStudent Found")
            print("Name :", student["name"])
            print("Marks:", student["marks"])

            found = True
            break

    if not found:
        print("Student not found.")


def display_above_75(students):

    print("\nStudents Above 75 Marks")

    for student in students:

        if student["marks"] > 75:
            print(student["name"], "-", student["marks"])


def display_below_35(students):

    print("\nStudents Below 35 Marks")

    for student in students:

        if student["marks"] < 35:
            print(student["name"], "-", student["marks"])


def create_summary(stats):

    wb = load_workbook(FILE_PATH)

    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws = wb.create_sheet("Summary")

    ws.append(["Metric", "Value"])

    for key, value in stats.items():
        ws.append([key, value])

    wb.save(FILE_PATH)

    print("\nSummary sheet created successfully.")

def main():

    students = read_students()

    if not students:
        return

    stats, above_75, below_35 = generate_statistics(students)

    print("\n----- STATISTICS -----")

    for key, value in stats.items():
        print(key, ":", value)

    print("\nStudents Above 75 Marks")

    for student in above_75:
        print(student["name"], "-", student["marks"])

    print("\nStudents Below 35 Marks")

    for student in below_35:
        print(student["name"], "-", student["marks"])

    search_student(students)

    create_summary(stats)

main()